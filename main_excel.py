# -------------------------------------------------------------------------------
# NOT USED - old main, used for executing scenarios, and saving results of each run of scenario to excel
# replaced by main_shelve
# -------------------------------------------------------------------------------

import simulation
import time
import openpyxl
import scenario as sc
from math import trunc
import mytimes
import global_settings
import pickle
import multiprocessing as mp
from main_shelve import generate_demands


def print_results_to_sheet(results, sheet, offset_row, start_column):
    for i in range(len(results)):
        for j in range(len(results[i])):
            sheet.cell(row=first_row + offset_row, column=start_column + 3 * j + i, value=round(results[i][j], 2))


def print_times():
    round_by = 3
    summ = 0
    sums = []
    for num, i in enumerate(mytimes.exec_groups):
        print(num, [round(x, round_by) for x in i])
        for j in i:
            summ += j
    for i in zip(*mytimes.exec_groups):
        sums.append(sum(i))
    sums = [round(i, round_by) for i in sums]
    print("\nS", sums)
    print("total sum: ", round(summ, round_by))
    print("\n")


# todo: was nehmen für holding costs warehouse?

# muss noch rein in meth. IN_t = IN_t-1  + O_t - mu_t

# josef: lead time = 0 auch als case? erstmal nicht wichtig, L=2 ist okay! Q eoq auch gut
#    retailer mit unterschiedichen lead times auch als case? wenn ja, wie wird dann vorausgerechnet? kriegt jeder retailer sein eigenes zeitfenster(->ja)?
# einleitung, literatur,
# rückwärtssuche (für aktuelle papers zum thema): wer zitiert zB. gallego2007?, axsäter

first_row = 4

def run_scenario(scenario):
    if scenario.duration < 100:
        temp_name = 'short'
    elif scenario.duration < 1000:
        temp_name = 'mid'
    elif scenario.duration < 10000:
        temp_name = 'long'
    else:
        temp_name = 'reeeally long'

    wb = openpyxl.load_workbook(
        './generated_sheets/templates/template ' + temp_name + '.xlsx',
        read_only=False)
    sheet = wb[wb.sheetnames[0]]

    r = scenario.get_iterator()
    for current in r:
        sim = simulation.Simulation(length=scenario.length, warm_up=scenario.warm_up, stock=60,
                                    high_var=scenario.high_var,
                                    high_c_shortage=scenario.high_c_shortage, demands=scenario.demands,
                                    distribution=scenario.distribution, L0=scenario.L0)
        print(current)
        print(round(((current[3] / scenario.duration) * 100), 2), "%")

        sim.warehouse.R = current[0]
        sim.warehouse.retailers[0].R = current[1]
        sim.warehouse.retailers[1].R = current[2]

        sheet["A%d" % (first_row + current[3])] = sim.warehouse.R
        sheet["B%d" % (first_row + current[3])] = sim.warehouse.retailers[0].R
        sheet["C%d" % (first_row + current[3])] = sim.warehouse.retailers[1].R
        sheet["D%d" % (first_row + current[3])] = str(sim.warehouse.retailers[0].seed) + "," + str(
            sim.warehouse.retailers[1].seed)

        global_settings.full_batches = False
        pre1 = time.time()
        sim.run(FIFO=False)
        after1 = time.time()
        print_results_to_sheet(sim.collect_statistics(), sheet, current[3], 8)
        sim.reset()

        global_settings.full_batches = True
        pre2 = time.time()
        sim.run(FIFO=False)
        after2 = time.time()
        print_results_to_sheet(sim.collect_statistics(), sheet, current[3], 19)
        sim.reset()

        pre3 = time.time()
        sim.run(FIFO=True)
        after3 = time.time()
        print_results_to_sheet(sim.collect_statistics(), sheet, current[3], 31)
        sim.reset()

        sheet["E%d" % (first_row + current[3])] = after1 - pre1
        sheet["F%d" % (first_row + current[3])] = after2 - pre2
        sheet["G%d" % (first_row + current[3])] = after3 - pre3

    sheet["AW4"] = scenario.length - scenario.warm_up
    sheet["AW8"] = sim.distribution[0]
    sheet["AX8"] = sim.distribution[1]
    sheet["AY8"] = float(sim.distribution[2])
    sheet["AZ8"] = float(sim.distribution[3])
    sheet["BA8"] = float(sim.distribution[4])

    sheet["AW12"] = sim.warehouse.retailers[0].c_holding
    sheet["AX12"] = sim.warehouse.retailers[0].c_shortage
    sheet["AY12"] = sim.warehouse.retailers[1].c_holding
    sheet["AZ12"] = sim.warehouse.retailers[1].c_shortage

    name = "generated_sheets/" + str(scenario.name) + ".xlsx"
    wb.save(name)


if __name__ == '__main__':
    periods = 10000
    warm_up = 100
    high_var = True

    demands_high, distribution_high = generate_demands(periods + warm_up, True)
    demands_low, distribution_low = generate_demands(periods + warm_up, False)

    #initially generated demands with generate_demands() and saved them - now, they are loaded using pickle
    with open("demands_high.txt", "rb") as f:
        demands_high = pickle.load(f)
    with open("demands_low.txt", "rb") as f:
        demands_low = pickle.load(f)

    r1, r2, r3 = (15, 60), (25, 65), (25, 65)
    settings1 = {"high_c_shortage": True, "L0": 2}
    s1 = sc.Scenario("equal retailers test", periods, warm_up, r1, r2, r3, 15, 1,
                     1,
                     repeat=1,
                     high_var=True, run_me_as=0, demands=demands_high,
                     distribution=distribution_high, settings=settings1)
    run_scenario(s1)
